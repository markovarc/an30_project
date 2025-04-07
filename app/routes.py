# app/routes.py

import sqlite3
from flask import render_template, request, redirect, send_file, url_for
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from . import app
from .database import get_db, init_db, get_next_free_id

# Сколько записей на одной странице (для пагинации)
RECORDS_PER_PAGE = 10

# Цвета, используемые в приложении
COLORS = {
    'primary': "#6C7A89",
    'secondary': "#95A5A6",
    'background': "#F5F7FA",
    'accent': "#4A90E2",
    'danger': "#ff4444",
    'status': {
        'work':    "#C8E6C9",
        'stop':    "#FFCDD2",
        'repair':  "#FFF9C4",
        'holiday': "#E1BEE7"
    }
}

# --------------------- ФУНКЦИИ ДОБАВЛЕНИЯ В БД ---------------------

def insert_machine(name: str):
    conn = get_db()
    try:
        new_id = get_next_free_id(conn, "machines")
        conn.execute("INSERT INTO machines (id, name) VALUES (?,?)", (new_id, name))
        conn.commit()
    finally:
        conn.close()

def insert_driver(name: str):
    conn = get_db()
    try:
        new_id = get_next_free_id(conn, "drivers")
        conn.execute("INSERT INTO drivers (id, name) VALUES (?,?)", (new_id, name))
        conn.commit()
    finally:
        conn.close()

def insert_counterparty(name: str):
    conn = get_db()
    try:
        new_id = get_next_free_id(conn, "counterparties")
        conn.execute("INSERT INTO counterparties (id, name) VALUES (?,?)", (new_id, name))
        conn.commit()
    finally:
        conn.close()

def insert_record(date_str, machine_id, driver_id, status, start_t, end_t, hours, comment, cpar_id):
    conn = get_db()
    try:
        new_id = get_next_free_id(conn, "records")
        conn.execute('''
            INSERT INTO records
            (id, date, machine_id, driver_id, status, start_time, end_time, hours, comment, counterparty_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (new_id, date_str, machine_id, driver_id, status, start_t, end_t, hours, comment, cpar_id))
        conn.commit()
    finally:
        conn.close()

# --------------------- ГЛАВНАЯ СТРАНИЦА (список техники) ---------------------

@app.route('/')
def index():
    conn = get_db()
    try:
        machines = conn.execute("SELECT * FROM machines ORDER BY id").fetchall()
    finally:
        conn.close()

    return render_template('index.html', machines=machines)

# --------------------- КАЛЕНДАРЬ ---------------------

@app.route('/calendar/<int:machine_id>')
def calendar(machine_id):
    year = request.args.get('year', type=int, default=datetime.now().year)
    month = request.args.get('month', type=int, default=datetime.now().month)
    if month < 1: 
        month = 1
    if month > 12: 
        month = 12
    if year < 2020: 
        year = 2020
    if year > 2030: 
        year = 2030

    conn = get_db()
    try:
        machine = conn.execute("SELECT * FROM machines WHERE id=?", (machine_id,)).fetchone()
        if not machine:
            # Если техника не найдена
            return render_template('base.html', content="<h2>Такой техники нет</h2>"), 404

        first_day = datetime(year, month, 1)
        last_day  = (first_day.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        dates = [first_day + timedelta(days=i) for i in range((last_day - first_day).days + 1)]

        # Получаем записи для каждого дня
        recs_dict = {}
        for d in dates:
            day_recs = conn.execute('''
                SELECT 
                  CASE 
                    WHEN r.driver_id IS NULL THEN '—'
                    WHEN d.id IS NULL THEN 'Водитель удалён'
                    ELSE d.name
                  END AS driver_name,
                  r.status,
                  r.start_time,
                  r.end_time,
                  CASE
                    WHEN r.counterparty_id IS NULL THEN ''
                    WHEN c.id IS NULL THEN 'Контрагент удалён'
                    ELSE c.name
                  END AS counterparty_name
                FROM records r
                LEFT JOIN drivers d ON r.driver_id=d.id
                LEFT JOIN counterparties c ON r.counterparty_id=c.id
                WHERE r.machine_id=? AND r.date=?
            ''', (machine_id, d.date())).fetchall()
            recs_dict[d.date()] = day_recs
    finally:
        conn.close()

    # Рассчитываем ссылки на предыдущий/следующий месяц
    prev_month = month - 1
    prev_year = year
    if prev_month < 1:
        prev_month = 12
        prev_year -= 1

    next_month = month + 1
    next_year = year
    if next_month > 12:
        next_month = 1
        next_year += 1

    return render_template('calendar.html',
                           machine=machine,
                           year=year,
                           month=month,
                           dates=dates,
                           recs_dict=recs_dict,
                           COLORS=COLORS,
                           prev_year=prev_year,
                           prev_month=prev_month,
                           next_year=next_year,
                           next_month=next_month)

# --------------------- АДМИНКА (меню) ---------------------

@app.route('/admin')
def admin():
    return render_template('admin.html')

# --------------------- СПРАВОЧНИКИ: МАШИНЫ ---------------------

@app.route('/admin/machines', methods=['GET','POST'])
def admin_machines():
    if request.method == 'POST':
        insert_machine(request.form['name'])
        return redirect('/admin/machines')

    conn = get_db()
    try:
        machines = conn.execute("SELECT * FROM machines ORDER BY id").fetchall()
    finally:
        conn.close()

    return render_template('admin_machines.html', machines=machines)

@app.route('/edit/machine/<int:id>', methods=['GET','POST'])
def edit_machine(id):
    conn = get_db()
    if request.method == 'POST':
        new_name = request.form['name']
        try:
            conn.execute("UPDATE machines SET name=? WHERE id=?", (new_name, id))
            conn.commit()
        except:
            conn.rollback()
        finally:
            conn.close()
        return redirect('/admin/machines')
    else:
        machine = conn.execute("SELECT * FROM machines WHERE id=?", (id,)).fetchone()
        conn.close()
        if not machine:
            return render_template('base.html', content="<h2>Машина не найдена</h2>"), 404
        return render_template('admin_machines.html', machine_edit=machine)

@app.route('/delete/machine/<int:id>', methods=['POST'])
def delete_machine(id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM machines WHERE id=?", (id,))
        conn.commit()
    except:
        conn.rollback()
        return "Ошибка удаления", 500
    finally:
        conn.close()
    return redirect('/admin/machines')

# --------------------- СПРАВОЧНИКИ: ВОДИТЕЛИ ---------------------

@app.route('/admin/drivers', methods=['GET','POST'])
def admin_drivers():
    if request.method == 'POST':
        insert_driver(request.form['name'])
        return redirect('/admin/drivers')

    conn = get_db()
    try:
        drivers = conn.execute("SELECT * FROM drivers ORDER BY id").fetchall()
    finally:
        conn.close()

    return render_template('admin_drivers.html', drivers=drivers)

@app.route('/edit/driver/<int:id>', methods=['GET','POST'])
def edit_driver(id):
    conn = get_db()
    if request.method == 'POST':
        new_name = request.form['name']
        try:
            conn.execute("UPDATE drivers SET name=? WHERE id=?", (new_name, id))
            conn.commit()
        except:
            conn.rollback()
        finally:
            conn.close()
        return redirect('/admin/drivers')
    else:
        dr = conn.execute("SELECT * FROM drivers WHERE id=?", (id,)).fetchone()
        conn.close()
        if not dr:
            return render_template('base.html', content="<h2>Водитель не найден</h2>"), 404
        return render_template('admin_drivers.html', driver_edit=dr)

@app.route('/delete/driver/<int:id>', methods=['POST'])
def delete_driver(id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM drivers WHERE id=?", (id,))
        conn.commit()
    except:
        conn.rollback()
        return "Ошибка удаления", 500
    finally:
        conn.close()
    return redirect('/admin/drivers')

# --------------------- СПРАВОЧНИКИ: КОНТРАГЕНТЫ ---------------------

@app.route('/admin/counterparties', methods=['GET','POST'])
def admin_counterparties():
    if request.method == 'POST':
        insert_counterparty(request.form['name'])
        return redirect('/admin/counterparties')

    conn = get_db()
    try:
        cparties = conn.execute("SELECT * FROM counterparties ORDER BY id").fetchall()
    finally:
        conn.close()

    return render_template('admin_counterparties.html', cparties=cparties)

@app.route('/edit/counterparty/<int:id>', methods=['GET','POST'])
def edit_counterparty(id):
    conn = get_db()
    if request.method == 'POST':
        new_name = request.form['name']
        try:
            conn.execute("UPDATE counterparties SET name=? WHERE id=?", (new_name, id))
            conn.commit()
        except:
            conn.rollback()
        finally:
            conn.close()
        return redirect('/admin/counterparties')
    else:
        cp = conn.execute("SELECT * FROM counterparties WHERE id=?", (id,)).fetchone()
        conn.close()
        if not cp:
            return render_template('base.html', content="<h2>Контрагент не найден</h2>"), 404
        return render_template('admin_counterparties.html', cparty_edit=cp)

@app.route('/delete/counterparty/<int:id>', methods=['POST'])
def delete_counterparty(id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM counterparties WHERE id=?", (id,))
        conn.commit()
    except:
        conn.rollback()
        return "Ошибка удаления", 500
    finally:
        conn.close()
    return redirect('/admin/counterparties')

# --------------------- СПИСОК ЗАПИСЕЙ (records) ---------------------

@app.route('/admin/records', methods=['GET','POST'])
def admin_records():
    if request.method == 'POST':
        # Добавляем запись
        date_str  = request.form['date']
        machine_id = int(request.form['machine_id'])
        driver_id  = int(request.form['driver_id'])
        status     = request.form['status']
        st_time    = request.form.get('start_time','')
        end_time   = request.form.get('end_time','')
        comm       = request.form.get('comment','')
        c_id       = request.form.get('counterparty_id')
        cpar_id    = int(c_id) if c_id else None

        hours = 0
        if st_time and end_time:
            try:
                st = datetime.strptime(st_time, '%H:%M')
                en = datetime.strptime(end_time, '%H:%M')
                if en < st:
                    en += timedelta(days=1)
                delta = en - st
                hours = delta.seconds // 3600
            except:
                pass

        insert_record(date_str, machine_id, driver_id, status, 
                      st_time or None, end_time or None,
                      hours, comm, cpar_id)
        return redirect('/admin/records')

    # Фильтры
    date_from = request.args.get('date_from','')
    date_to   = request.args.get('date_to','')
    mach_f    = request.args.get('mach', type=int)
    driv_f    = request.args.get('driv', type=int)
    cpar_f    = request.args.get('cpar', type=int)
    stat_f    = request.args.get('status','')
    comm_sub  = request.args.get('comment_sub','').strip()
    sort_key  = request.args.get('sort','date_desc')
    page      = request.args.get('page', type=int, default=1)
    if page < 1:
        page = 1

    where = []
    pr = []
    if date_from:
        where.append("r.date>=?")
        pr.append(date_from)
    if date_to:
        where.append("r.date<=?")
        pr.append(date_to)
    if mach_f:
        where.append("r.machine_id=?")
        pr.append(mach_f)
    if driv_f:
        where.append("r.driver_id=?")
        pr.append(driv_f)
    if cpar_f:
        where.append("r.counterparty_id=?")
        pr.append(cpar_f)
    if stat_f in ("work","stop","repair","holiday"):
        where.append("r.status=?")
        pr.append(stat_f)
    if comm_sub:
        where.append("r.comment LIKE ?")
        pr.append(f"%{comm_sub}%")

    where_sql = ""
    if where:
        where_sql = "WHERE " + " AND ".join(where)

    if sort_key=="date_asc":
        order_sql="ORDER BY r.date ASC, r.id ASC"
    elif sort_key=="date_desc":
        order_sql="ORDER BY r.date DESC, r.id DESC"
    elif sort_key=="hours_asc":
        order_sql="ORDER BY r.hours ASC, r.date ASC"
    elif sort_key=="hours_desc":
        order_sql="ORDER BY r.hours DESC, r.date DESC"
    elif sort_key=="machine_asc":
        order_sql="ORDER BY machine_name ASC, r.date DESC"
    elif sort_key=="driver_asc":
        order_sql="ORDER BY driver_name ASC, r.date DESC"
    else:
        order_sql="ORDER BY r.date DESC, r.id DESC"

    conn = get_db()
    count_sql = f'''
        SELECT COUNT(*)
          FROM records r
          LEFT JOIN machines m ON r.machine_id=m.id
          LEFT JOIN drivers d ON r.driver_id=d.id
          LEFT JOIN counterparties c ON r.counterparty_id=c.id
        {where_sql}
    '''
    total_count = conn.execute(count_sql, pr).fetchone()[0]
    offset = (page-1)*RECORDS_PER_PAGE
    total_pages = (total_count+RECORDS_PER_PAGE-1)//RECORDS_PER_PAGE

    sql = f'''
    SELECT 
      r.id,
      r.date,
      CASE 
        WHEN r.machine_id IS NULL THEN '—'
        WHEN m.id IS NULL THEN 'Техника удалена'
        ELSE m.name
      END AS machine_name,
      CASE
        WHEN r.driver_id IS NULL THEN '—'
        WHEN d.id IS NULL THEN 'Водитель удалён'
        ELSE d.name
      END AS driver_name,
      r.start_time,
      r.end_time,
      r.hours,
      r.comment,
      CASE
        WHEN r.counterparty_id IS NULL THEN '—'
        WHEN c.id IS NULL THEN 'Контрагент удалён'
        ELSE c.name
      END AS cparty_name,
      r.status
    FROM records r
    LEFT JOIN machines m ON r.machine_id=m.id
    LEFT JOIN drivers  d ON r.driver_id=d.id
    LEFT JOIN counterparties c ON r.counterparty_id=c.id
    {where_sql}
    {order_sql}
    LIMIT ? OFFSET ?
    '''
    records = conn.execute(sql, pr + [RECORDS_PER_PAGE, offset]).fetchall()

    machines = conn.execute("SELECT * FROM machines ORDER BY id").fetchall()
    drivers  = conn.execute("SELECT * FROM drivers ORDER BY id").fetchall()
    cparties = conn.execute("SELECT * FROM counterparties ORDER BY id").fetchall()
    conn.close()

    return render_template('admin_records.html',
                           records=records,
                           machines=machines,
                           drivers=drivers,
                           cparties=cparties,
                           COLORS=COLORS,
                           date_from=date_from,
                           date_to=date_to,
                           mach_f=mach_f,
                           driv_f=driv_f,
                           cpar_f=cpar_f,
                           stat_f=stat_f,
                           comm_sub=comm_sub,
                           sort_key=sort_key,
                           page=page,
                           total_pages=total_pages,
                           total_count=total_count,
                           RECORDS_PER_PAGE=RECORDS_PER_PAGE)

@app.route('/edit/record/<int:id>', methods=['GET','POST'])
def edit_record(id):
    conn = get_db()
    if request.method == 'POST':
        try:
            date_str  = request.form['date']
            machine_id = int(request.form['machine_id'])
            driver_id  = int(request.form['driver_id'])
            status     = request.form['status']
            st_t       = request.form.get('start_time','')
            end_t      = request.form.get('end_time','')
            comm       = request.form.get('comment','')
            c_id       = request.form.get('counterparty_id')
            cpar_id    = int(c_id) if c_id else None

            hrs = 0
            if st_t and end_t:
                try:
                    sdt = datetime.strptime(st_t, '%H:%M')
                    edt = datetime.strptime(end_t, '%H:%M')
                    if edt < sdt:
                        edt += timedelta(days=1)
                    delta = edt - sdt
                    hrs = delta.seconds // 3600
                except:
                    pass

            conn.execute('''
                UPDATE records
                   SET date=?,
                       machine_id=?,
                       driver_id=?,
                       status=?,
                       start_time=?,
                       end_time=?,
                       hours=?,
                       comment=?,
                       counterparty_id=?
                 WHERE id=?
            ''', (date_str, machine_id, driver_id, status, 
                  st_t or None, end_t or None, hrs, comm, cpar_id, id))
            conn.commit()
        except Exception as e:
            print(f"Ошибка редактирования записи: {e}")
            conn.rollback()
        finally:
            conn.close()
        return redirect('/admin/records')
    else:
        record = conn.execute('''
            SELECT date, machine_id, driver_id, status, start_time, end_time, hours, comment, counterparty_id
              FROM records
             WHERE id=?
        ''', (id,)).fetchone()

        machines = conn.execute("SELECT * FROM machines ORDER BY id").fetchall()
        drivers  = conn.execute("SELECT * FROM drivers ORDER BY id").fetchall()
        cparties = conn.execute("SELECT * FROM counterparties ORDER BY id").fetchall()
        conn.close()

        if not record:
            return render_template('base.html', content="<h2>Запись не найдена</h2>"), 404

        return render_template('record_edit.html',
                               rec_id=id,
                               record=record,
                               machines=machines,
                               drivers=drivers,
                               cparties=cparties)

@app.route('/delete/record/<int:id>', methods=['POST'])
def delete_record(id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM records WHERE id=?", (id,))
        conn.commit()
    except:
        conn.rollback()
        return "Ошибка удаления записи", 500
    finally:
        conn.close()
    return redirect('/admin/records')

# --------------------- ЭКСПОРТ В EXCEL ---------------------

@app.route('/export')
def export_excel():
    export_mode = request.args.get('export','')
    conn = get_db()
    try:
        if export_mode == 'filtered':
            # Читаем те же фильтры
            date_from = request.args.get('date_from','')
            date_to   = request.args.get('date_to','')
            mach_f    = request.args.get('mach', type=int)
            driv_f    = request.args.get('driv', type=int)
            cpar_f    = request.args.get('cpar', type=int)
            stat_f    = request.args.get('status','')
            comm_sub  = request.args.get('comment_sub','').strip()
            sort_key  = request.args.get('sort','date_desc')

            where = []
            pr = []
            if date_from:
                where.append("r.date>=?")
                pr.append(date_from)
            if date_to:
                where.append("r.date<=?")
                pr.append(date_to)
            if mach_f:
                where.append("r.machine_id=?")
                pr.append(mach_f)
            if driv_f:
                where.append("r.driver_id=?")
                pr.append(driv_f)
            if cpar_f:
                where.append("r.counterparty_id=?")
                pr.append(cpar_f)
            if stat_f in ("work","stop","repair","holiday"):
                where.append("r.status=?")
                pr.append(stat_f)
            if comm_sub:
                where.append("r.comment LIKE ?")
                pr.append(f"%{comm_sub}%")

            where_sql = ""
            if where:
                where_sql = "WHERE " + " AND ".join(where)

            if sort_key=="date_asc":
                order_sql="ORDER BY r.date ASC, r.id ASC"
            elif sort_key=="date_desc":
                order_sql="ORDER BY r.date DESC, r.id DESC"
            elif sort_key=="hours_asc":
                order_sql="ORDER BY r.hours ASC, r.date ASC"
            elif sort_key=="hours_desc":
                order_sql="ORDER BY r.hours DESC, r.date DESC"
            elif sort_key=="machine_asc":
                order_sql="ORDER BY machine_name ASC, r.date DESC"
            elif sort_key=="driver_asc":
                order_sql="ORDER BY driver_name ASC, r.date DESC"
            else:
                order_sql="ORDER BY r.date DESC, r.id DESC"

            rows = conn.execute(f'''
                SELECT 
                  r.date,
                  CASE
                    WHEN r.machine_id IS NULL THEN '—'
                    WHEN m.id IS NULL THEN 'Техника удалёна'
                    ELSE m.name
                  END AS machine_name,
                  CASE
                    WHEN r.driver_id IS NULL THEN '—'
                    WHEN d.id IS NULL THEN 'Водитель удалён'
                    ELSE d.name
                  END AS driver_name,
                  r.status,
                  IFNULL(r.start_time,""),
                  IFNULL(r.end_time,""),
                  r.hours,
                  CASE
                    WHEN r.counterparty_id IS NULL THEN '—'
                    WHEN c.id IS NULL THEN 'Контрагент удалён'
                    ELSE c.name
                  END AS cparty_name,
                  IFNULL(r.comment,"-")
                FROM records r
                LEFT JOIN machines m ON r.machine_id=m.id
                LEFT JOIN drivers d  ON r.driver_id=d.id
                LEFT JOIN counterparties c ON r.counterparty_id=c.id
                {where_sql}
                {order_sql}
            ''', pr).fetchall()
        else:
            # Экспорт всех записей
            rows = conn.execute('''
                SELECT
                  r.date,
                  CASE
                    WHEN r.machine_id IS NULL THEN '—'
                    WHEN m.id IS NULL THEN 'Техника удалёна'
                    ELSE m.name
                  END AS machine_name,
                  CASE
                    WHEN r.driver_id IS NULL THEN '—'
                    WHEN d.id IS NULL THEN 'Водитель удалён'
                    ELSE d.name
                  END AS driver_name,
                  r.status,
                  IFNULL(r.start_time,""),
                  IFNULL(r.end_time,""),
                  r.hours,
                  CASE
                    WHEN r.counterparty_id IS NULL THEN '—'
                    WHEN c.id IS NULL THEN 'Контрагент удалён'
                    ELSE c.name
                  END AS cparty_name,
                  IFNULL(r.comment,"-")
                FROM records r
                LEFT JOIN machines m ON r.machine_id=m.id
                LEFT JOIN drivers d ON r.driver_id=d.id
                LEFT JOIN counterparties c ON r.counterparty_id=c.id
                ORDER BY r.date ASC, r.id ASC
            ''').fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "AN-30 Отчёт"

    headers = ["Дата", "Техника", "Водитель", "Статус", "Начало", "Конец", "Часы", "Контрагент", "Комментарий"]
    ws.append(headers)

    header_fill = PatternFill(start_color="444444", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col)].width = 20

    for row in rows:
        date_db     = row[0]
        machine_nm  = row[1]
        driver_nm   = row[2]
        status_     = row[3]
        start_      = row[4]
        end_        = row[5]
        hrs_        = row[6] or 0
        cpar_       = row[7]
        comm_       = row[8]

        # Перевод даты в формат дд.мм.гггг
        try:
            date_fmt = datetime.strptime(date_db, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_fmt = date_db

        st_hex = COLORS['status'].get(status_, "#FFFFFF")[1:]  # берем цвет без "#"

        ws.append([
            date_fmt, machine_nm, driver_nm, 
            status_.capitalize(), start_, end_, hrs_, cpar_, comm_
        ])
        # Заливаем ячейку статуса
        st_cell = ws.cell(row=ws.max_row, column=4)
        st_cell.fill = PatternFill(start_color=st_hex, fill_type="solid")

    fname = "report_" + datetime.now().strftime("%Y%m%d_%H%M") + ".xlsx"
    wb.save(fname)

    return send_file(
        fname,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
